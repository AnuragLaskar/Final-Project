import re
from openpyxl import load_workbook
wb = load_workbook(filename = 'abe.xlsx')

print wb.get_sheet_names()
sh = wb.get_sheet_names()
ws = wb[sh[0]]

file = open("cleandata.txt","w")

def clean_tweet(tweet):
	return ' '.join(re.sub("(@[A-Za-z0-9]+)|([^0-9A-Za-z \t]) |(\w+:\/\/\S+)", " ", tweet).split())

tweets = []

for i in ws['D']:
    if (i.value) is not None:
        if(clean_tweet(i.value.encode("utf-8"))) not in tweets:
            tweets.append(clean_tweet(i.value.encode("utf-8")))
            file.write(clean_tweet(i.value.encode("utf-8")))
            file.write("\n")
file.close()